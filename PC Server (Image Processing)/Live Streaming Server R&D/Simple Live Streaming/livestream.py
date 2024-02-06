# Ref : https://www.bogotobogo.com/python/OpenCV_Python/python_opencv3_Image_Object_Detection_Face_Detection_Haar_Cascade_Classifiers.php
from threading import Thread
from os.path import dirname, abspath
import os
import cv2
import urllib.request
import numpy as np
import time
import openpyxl
import requests

common_resources_dir = "\Common Resources"
root_folder_content_signature = common_resources_dir  # To find root folder automatically

# EXCEL IP CAMS DATABASE
IPCAMS_DATABASE = common_resources_dir + '\All_IPCams.xlsx'
BAT_ID = 1
IPCAMS_DATABASE_STARTROW = 3
IPCAMS_DATABASE_STARTCOL = 6

# CAMS
W_DEFAULT, H_DEFAULT = 640, 480
h, w = W_DEFAULT, H_DEFAULT
are_frontals = True

################## Get root project directory ##################
i = 0
root_project_dir = dirname(abspath(__file__))
while not os.path.exists(root_project_dir + root_folder_content_signature):
    root_project_dir = dirname(abspath(root_project_dir))
    print(root_project_dir)
    i += 1
    assert i < 100, f'Root project directory not found. Please ensure that there is a "\{root_folder_content_signature}" folder inside it'

root_project_dir += "\""
print(f'Root project directory found : \n {root_project_dir}')
#################################################################




##############  Init cameras video stream ###############
# Get Cams Urls
IPCAMS_DATABASE_PATH = root_project_dir + IPCAMS_DATABASE
IPCAMS_DATABASE_PATH = IPCAMS_DATABASE_PATH.replace("\"", "\\")  # Convert to double backslashes if needed
wb = openpyxl.load_workbook(IPCAMS_DATABASE_PATH, data_only=True)
bat_names = wb.sheetnames
ws = wb[bat_names[BAT_ID]]
urls = [ws.cell(row=i + IPCAMS_DATABASE_STARTROW, column=IPCAMS_DATABASE_STARTCOL).value for i in
        range(1, ws.max_row + 1) if
        ws.cell(row=i + IPCAMS_DATABASE_STARTROW, column=IPCAMS_DATABASE_STARTCOL).value is not None]
print(urls)
n = len(urls)

# Prepare imgs
imgs = [np.zeros((H_DEFAULT, W_DEFAULT, 3)) for i in urls]

###########################################################

global img


def url_ok(url, timeout):
    try:
        r = requests.head(url, timeout=timeout)
        return r.status_code == 200
    except:
        return False


def update(index, url):
    # Read next stream frame in a daemon thread
    global w, h
    f = 0
    img_cv = None
    while True:
        f += 1
        success = False
        if f == 1:
            # Meth 1 for getting image from url
            try:
                if (url_ok(url, 1)):
                    img_rawInput = urllib.request.urlopen(url)
                    img_numpy = np.array(bytearray(img_rawInput.read()), dtype=np.uint8)
                    img_cv = cv2.imdecode(img_numpy, -1)
                    # Orientation correction
                    img_cv = cv2.transpose(img_cv)
                    if not are_frontals:
                        img_cv = cv2.flip(img_cv, 1)

                    success = True
            except:
                print(f'failed to open cam n°{index:d}')

            # Meth 2 for getting image from url
            # try:
            #     if (url_ok(url, 1)):
            #         cap = cv2.VideoCapture(url)
            #         success, im = cap.read()
            # except:
            #     print(f'failed to open cam no : {index:d}')

            if img_cv is None:
                img_cv = np.zeros((h, w, 3))
            else:
                h, w = img_cv.shape[0], img_cv.shape[1]
            imgs[index] = img_cv
            #time.sleep(1 / 25)  # wait time
            f = 0


########## Prepare camera threads #########
for i, url in enumerate(urls):
    print(url)
    thread = Thread(target=update, args=([i, url]), daemon=True)
    thread.start()

########## Display live streams #########
while True:
    for i, img_out in enumerate(imgs):
        cv2.imshow(f'Live cam testing : {i + 1:d}/{n:d} ({urls[i]})', img_out)
    key = cv2.waitKey(1)
    if key == ord('q'):
        break

cv2.destroyAllWindows()
