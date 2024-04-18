# Ref : https://www.bogotobogo.com/python/OpenCV_Python/python_opencv3_Image_Object_Detection_Face_Detection_Haar_Cascade_Classifiers.php
from os.path import dirname, abspath
import cv2
import numpy as np
import openpyxl
import argparse
import matplotlib.pyplot as plt
import time

from utils.plots import colors, plot_one_box_kpt, plot_human
from camera import GenericCamera
from CameraSync import Camera
from poseestimation.poseestimator import PoseEstimator
from poseclassification.poseclassifier import SymbolicPoseClassifier
from poseclassification.humanExtractor import HumanExtractor
from mathutils.vectors import*
from multiview.BiCamera import BiCamera
from multiview.CamGroup import CamGroup


from network.localnetwork import *

common_resources_dir = "\#Common Resources"
root_folder_content_signature = common_resources_dir  # To find root folder automatically

# EXCEL IP CAMS DATABASE
IPCAMS_DATABASE = common_resources_dir + '\All_IPCams.xlsx'
BAT_ID = 1
IPCAMS_DATABASE_STARTROW = 3
IPCAMS_DATABASE_STARTCOL = 6

# CAMS
W_DEFAULT, H_DEFAULT = 800, 600  # 768, 576
w, h = W_DEFAULT, H_DEFAULT

correls = []


def gcd_resize(input_frame, gcd):
    frame_height, frame_width = input_frame.shape[0], input_frame.shape[1]
    if gcd != 1:
        adapted_width = gcd * (frame_width // gcd)
        adapted_height = gcd * (frame_height // gcd)
        adapted_frame = cv2.resize(input_frame, (adapted_width, adapted_height))
    else:
        adapted_frame = input_frame
    return adapted_frame


def rotate_image(image, angle):
    # Get the image center
    height, width = image.shape[:2]
    center = (width // 2, height // 2)
    rotated_image = image

    if (angle != 0):
        # Calculate the rotation matrix
        rotation_matrix = cv2.getRotationMatrix2D(center, angle, 1.0)

        # Apply the rotation to the image
        rotated_image = cv2.warpAffine(image, rotation_matrix, (width, height))
    return rotated_image


########### Pose drawer ###############
def draw_poses(frame, detection_poses_results, names, line_thickness, hide_labels, hide_conf):
    for i, pose in enumerate(
            detection_poses_results):

        if len(detection_poses_results):  # check if no pose
            for c in pose[:, 5].unique():  # Print poses_results
                n = (pose[:, 5] == c).sum()  # detections per class
                print("No of Objects in Current Frame : {}".format(n))

            for det_index, (*xyxy, conf, cls) in enumerate(
                    reversed(pose[:, :6])):  # loop over poses for drawing on frame
                c = int(cls)  # integer class
                kpts = pose[det_index, 6:]
                label = None if hide_labels else (names[c] if hide_conf else f'{names[c]} {conf:.2f}')
                plot_one_box_kpt(xyxy, frame, label=label, color=colors(c, True),
                                 line_thickness=line_thickness, kpt_label=True, kpts=kpts, steps=3,
                                 orig_shape=frame.shape[:2])
    return frame


def ndarray_info(nda):
    print("type\t: ", nda.dtype, "\nshape\t: ", nda.shape, "\nmin\t: ", nda.min(), "\nmax\t: ", nda.max(), "\n")





class Main:

    def __init__(self, opt_arg):
        ################## Get root project directory (Only for development purposes) ##################
        i = 0
        root_project_dir = dirname(abspath(__file__))
        while not os.path.exists(root_project_dir + root_folder_content_signature):
            root_project_dir = dirname(abspath(root_project_dir))
            i += 1
            assert i < 100, f'Root project directory not found. Please ensure that there is a "\{root_folder_content_signature}" folder inside it'

        root_project_dir += "\""
        print(f'Root project directory found : \n {root_project_dir}')
        #################################################################

        ################# Get options ###############
        self.opt = opt_arg
        print(f"opt.source: {self.opt.source:s}")

        ############### Get LAN Infos ###############
        _, self.lan_ssid = get_current_lan_ssid()

        ##############  Init cameras video stream ###############
        self.cameras = []
        self.handlers=[]
        self.camgroups=[]

        if is_number(self.opt.source):
            source = int(self.opt.source)  # 0 : Test with PC webcam / !=0 : Use url cameras
            if source == 1:
                # Get Cams Urls
                IPCAMS_DATABASE_PATH = root_project_dir + IPCAMS_DATABASE
                IPCAMS_DATABASE_PATH = IPCAMS_DATABASE_PATH.replace("\"",
                                                                    "\\")  # Convert to double backslashes if needed
                wb = openpyxl.load_workbook(IPCAMS_DATABASE_PATH, data_only=True)
                bat_names = wb.sheetnames
                ws = wb[bat_names[BAT_ID]]
                self.urls = [ws.cell(row=i + IPCAMS_DATABASE_STARTROW, column=IPCAMS_DATABASE_STARTCOL).value for i in
                             range(1, ws.max_row + 1) if
                             ws.cell(row=i + IPCAMS_DATABASE_STARTROW,
                                     column=IPCAMS_DATABASE_STARTCOL).value is not None]
                # self.urls = self.urls[:1]
                print(self.urls)
                n = len(self.urls)
                rot_angles = [90 * 0 for i in range(n)]  # TODO rm 0
                rot_angles[3] = 0

                for i, url in enumerate(self.urls):
                    camera=Camera(url, rot_angles[i],(IPCAMS_DATABASE_STARTCOL == 4))
                    self.cameras.append(camera)

                for k in range(0,len(self.cameras),2):
                    camera1 = self.cameras[k]
                    camera1.status="master"
                    camera1.check_status()  # assignating the camera matrix
                    camera2 = self.cameras[k+1]
                    camera2.status="slave"
                    camera2.check_status()
                    myHandler=BiCamera(camera1, camera2)
                    myHandler.init()
                    myHandler.start()
                    self.handlers.append(myHandler)


            if source == 0:
                camera = GenericCamera(0, rot_angle=0)
                camera.source_is_auto_refresh = True
                camera.start()
                self.cameras.append(camera)

        else:  # Video file
            camera = GenericCamera(self.opt.source, rot_angle=0)
            camera.frame_time = 0.03
            camera.source_is_auto_refresh = True
            camera.start()
            self.cameras.append(camera)

        ###########################################################

        ################## Init pose estimator ###################
        self.pose_estimator = PoseEstimator(self.opt.device)
        self.names = self.pose_estimator.names
        ###########################################################

        ################## Init pose classifier ###################
        # Deprecated TODO rm
        self.pose_classifier = SymbolicPoseClassifier()
        self.humanExtractor = HumanExtractor(10, 852 - 10, 10, 480 - 10)

        ################# Creating CamGroups ######################
        for bicam in self.handlers:
            humanExtractorLeft=HumanExtractor(10, 852 - 10, 10, 480 - 10)
            humanExtractorRight=HumanExtractor(10, 852 - 10, 10, 480 - 10)

            camGroup=CamGroup(bicam, humanExtractorLeft, humanExtractorRight)
            self.camgroups.append(camGroup)

        ################# Pose association ##########################
        for camGroup in self.camgroups:
            left_frame= None #camGroup.biCam.  # TODO recuperer frame
            right_frame= None   # TODO recuperer frame

            output_frame_left, raw_results_left, keypoints_left, profiler_left = self.pose_estimator.detect(left_frame)
            output_frame_right, raw_results_right, keypoints_right, profiler_right = self.pose_estimator.detect(right_frame)

            camGroup.humanExtractorLeft.update_humans(raw_results_left, left_frame)
            camGroup.humanExtractorRight.update_humans(raw_results_right, right_frame)

        ################# Stereoscopic Correspondance ###############





        global start_time
        start_time = time.time()
        ###########################################################

        ################# Init output server ######################
        # TODO Fix problem here with output server init
        # self.out_stream = FrameStream()
        ###########################################################

    def update(self):
        global all_kpts
        j = 0
        n = len(self.cameras)
        output = [None for i in range(n)]

        for i, cam in enumerate(self.cameras):
            success, frame = cam.read()
            # Detection
            enable_detection = 1
            success = 1
            enable_pose_drawing = 1
            activity_classifier = 1
            if success:
                if enable_detection:
                    if i == j:
                        yol_frame, poses_results, _, _ = self.pose_estimator.detect(frame)
                        print("Current Frame shape : ", frame.shape)

                        if activity_classifier == 0:
                            dist, correl = self.pose_classifier.symbolic_classify(poses_results)
                            if enable_pose_drawing:
                                frame = draw_poses(frame, poses_results, self.names,
                                                   line_thickness=self.opt.line_thickness,
                                                   hide_conf=self.opt.hide_conf, hide_labels=self.opt.hide_labels)

                        if activity_classifier == 1:

                            for i, pose in enumerate(poses_results):

                                if len(poses_results):  # check if no pose
                                    for c in pose[:, 5].unique():  # Print poses_results
                                        n = (pose[:, 5] == c).sum()  # detections per class
                                        #print("No of Objects in Current Frame : {}".format(n))

                                        all_kpts = pose[:, 6:]

                                        for det_index, (*xyxy, conf, cls) in enumerate(
                                                reversed(pose[:, :6])):  # loop over poses for drawing on frame
                                            c = int(cls)  # integer class
                                            kpts = pose[det_index, 6:]

                                        # Pose classifier (Deprecated) TODO rm
                                        # current_time = time.time() - start_time
                                        # print(current_time)
                                        # print("Corrélation : " + str(correl.cpu().item()))

                                        # Human extractor
                                        self.humanExtractor.update_humans(all_kpts, yol_frame)
                                        humans = self.humanExtractor.humans

                                        current_time = time.time() - start_time

                                        if len(humans) > 1:
                                            correl_score = correlation(humans[0].get_pose(), humans[1].get_pose())
                                        print(current_time)
                                            #print("Correl between 1 and 3 :", correl_score.item())
                                            #print("Pose 1 :", humans[0].get_pose())
                                            #print("Pose 2 :", humans[1].get_pose())
                                            #print("Screen dims", frame.shape)

                                        for h, human in enumerate(humans):
                                            #print("humains sur scene : ", len(humans))
                                            human_kpts_xy = human.get_pose().reshape(-1, 3)[:, :-1]
                                            xyxy = [min(human_kpts_xy[:, 0]), min(human_kpts_xy[:, 1]), max(human_kpts_xy[:, 0]), max(human_kpts_xy[:, 1])]
                                            plot_human(human, frame, infos=True, state_text=("Normal" if h == 1 else "Malaise"), label=" humain:" + str(h + 1), color=colors(h, True),
                                                             line_thickness=6, kpt_label=True, steps=3,
                                                             orig_shape=frame.shape[:2])
                    output[i] = frame
                    frame = None

                    # print("Corrélation : " + str(correl.cpu().item()))

                    # try:
                    #     # Write data to a text file
                    #     if correl.cpu().item() > 0:
                    #         with open('correl.txt', 'a') as f:
                    #             f.write(f'{current_time:.2f} {correl.cpu().item()}\n')
                    # except Exception as e:
                    #     print('Error : %s', e)

            # Send the stream video
            # send_to_stream(frame, self.out_stream)

        j += 1
        j %= len(self.cameras)
        return output

    def close_cameras(self):
        for i, cam in enumerate(self.cameras):
            cam.stop()


def is_number(s):
    try:
        float(s)
        return True
    except ValueError:
        return False
