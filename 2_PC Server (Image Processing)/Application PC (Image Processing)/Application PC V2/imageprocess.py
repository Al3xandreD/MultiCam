# Ref : https://www.bogotobogo.com/python/OpenCV_Python/python_opencv3_Image_Object_Detection_Face_Detection_Haar_Cascade_Classifiers.php
from threading import Thread
from os.path import dirname, abspath

import os
import cv2
import urllib.request
import numpy as np
import time
import openpyxl
import requests
import argparse
import time
from pathlib import Path

import cv2
import torch
import torchvision
import torch.backends.cudnn as cudnn
from numpy import random

from models.experimental import attempt_load
from utils.datasets import LoadStreams, LoadImages
from utils.general import check_img_size, check_requirements, check_imshow, non_max_suppression, apply_classifier, \
    scale_coords, xyxy2xywh, strip_optimizer, set_logging, increment_path
from utils.plots import plot_one_box
from utils.torch_utils import select_device, load_classifier, time_synchronized, TracedModel

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--weights', nargs='+', type=str, default='yolov7.pt', help='model.pt path(s)')
    parser.add_argument('--source', type=str, default='inference/images', help='source')  # file/folder, 0 for webcam
    parser.add_argument('--img-size', type=int, default=640, help='inference size (pixels)')
    parser.add_argument('--conf-thres', type=float, default=0.25, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.45, help='IOU threshold for NMS')
    parser.add_argument('--device', default='', help='cuda device, i.e. 0 or 0,1,2,3 or cpu')
    parser.add_argument('--view-img', action='store_true', help='display results')
    parser.add_argument('--save-txt', action='store_true', help='save results to *.txt')
    parser.add_argument('--save-conf', action='store_true', help='save confidences in --save-txt labels')
    parser.add_argument('--nosave', action='store_true', help='do not save images/videos')
    parser.add_argument('--classes', nargs='+', type=int, help='filter by class: --class 0, or --class 0 2 3')
    parser.add_argument('--agnostic-nms', action='store_true', help='class-agnostic NMS')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    parser.add_argument('--update', action='store_true', help='update all models')
    parser.add_argument('--project', default='runs/detect', help='save results to project/name')
    parser.add_argument('--name', default='exp', help='save results to project/name')
    parser.add_argument('--exist-ok', action='store_true', help='existing project/name ok, do not increment')
    parser.add_argument('--no-trace', action='store_true', help='don`t trace model')
    opt = parser.parse_args()
    print(opt)
    #check_requirements(exclude=('pycocotools', 'thop'))

BODY_PARTS = {"Nose": 0, "Neck": 1, "RShoulder": 2, "RElbow": 3, "RWrist": 4,
              "LShoulder": 5, "LElbow": 6, "LWrist": 7, "RHip": 8, "RKnee": 9,
              "RAnkle": 10, "LHip": 11, "LKnee": 12, "LAnkle": 13, "REye": 14,
              "LEye": 15, "REar": 16, "LEar": 17, "Background": 18}

POSE_PAIRS = [["Neck", "RShoulder"], ["Neck", "LShoulder"], ["RShoulder", "RElbow"],
              ["RElbow", "RWrist"], ["LShoulder", "LElbow"], ["LElbow", "LWrist"],
              ["Neck", "RHip"], ["RHip", "RKnee"], ["RKnee", "RAnkle"], ["Neck", "LHip"],
              ["LHip", "LKnee"], ["LKnee", "LAnkle"], ["Neck", "Nose"], ["Nose", "REye"],
              ["REye", "REar"], ["Nose", "LEye"], ["LEye", "LEar"]]


def poseEstimation_run(frame):
    global frameHeight, frameWidth
    frameWidth = frame.shape[1]
    frameHeight = frame.shape[0]

    net.setInput(cv2.dnn.blobFromImage(frame, 1.0, (frameWidth, frameHeight), (127.5, 127.5, 127.5), swapRB=True, crop=False))
    global poseEstimation_out
    poseEstimation_out = net.forward()
    poseEstimation_out = poseEstimation_out[:, :19, :,
                         :]  # MobileNet output [1, 57, -1, -1], we only need the first 19 elements
    return (len(BODY_PARTS) == poseEstimation_out.shape[1])


def poseEstimation_postProcess(dst, offset):
    points = []
    for i in range(len(BODY_PARTS)):
        # Slice heatmap of corresponging body's part.
        heatMap = poseEstimation_out[0, i, :, :]

        # Originally, we try to find all the local maximums. To simplify a sample
        # we just find a global one. However only a single pose at the same time
        # could be detected this way.
        _, conf, _, point = cv2.minMaxLoc(heatMap)
        x = (frameWidth * point[0]) / poseEstimation_out.shape[3] + offset[0]
        y = (frameHeight * point[1]) / poseEstimation_out.shape[2] + offset[1]
        # Add a point if it's confidence is higher than threshold.
        points.append((int(x), int(y)) if conf > 0.2 else None)

    for pair in POSE_PAIRS:
        partFrom = pair[0]
        partTo = pair[1]
        assert (partFrom in BODY_PARTS)
        assert (partTo in BODY_PARTS)

        idFrom = BODY_PARTS[partFrom]
        idTo = BODY_PARTS[partTo]

        if points[idFrom] and points[idTo]:
            cv2.line(dst, points[idFrom], points[idTo], (0, 255, 0), 3)
            cv2.ellipse(dst, points[idFrom], (3, 3), 0, 0, 360, (0, 0, 255), cv2.FILLED)
            cv2.ellipse(dst, points[idTo], (3, 3), 0, 0, 360, (0, 0, 255), cv2.FILLED)


common_resources_dir = "\Common Resources"
root_folder_content_signature = common_resources_dir  # To find root folder automatically

# EXCEL IP CAMS DATABASE
IPCAMS_DATABASE = common_resources_dir + '\All_IPCams.xlsx'
BAT_ID = 1
BAT_NAMES = ['IOT', 'Rayane', 'Tamara']
IPCAMS_DATABASE_STARTROW = 3
IPCAMS_DATABASE_STARTCOL = 6

# CAMS
W_DEFAULT, H_DEFAULT = 640, 480
h, w = W_DEFAULT, H_DEFAULT
are_frontals = True

################## Get root project directory ##################
i = 0
root_project_dir = dirname(abspath(__file__))
while not os.path.exists(root_project_dir + root_folder_content_signature):
    root_project_dir = dirname(abspath(root_project_dir))
    print(root_project_dir)
    i += 1
    assert i < 100, f'Root project directory not found. Please ensure that there is a "\{root_folder_content_signature}" folder inside it'

root_project_dir += "\""
print(f'Root project directory found : \n {root_project_dir}')
#################################################################




#########  Init cameras video stream #########
# Get Cams Urls
IPCAMS_DATABASE_PATH = root_project_dir + IPCAMS_DATABASE
IPCAMS_DATABASE_PATH = IPCAMS_DATABASE_PATH.replace("\"", "\\")  # Convert to double backslashes if needed
wb = openpyxl.load_workbook(IPCAMS_DATABASE_PATH, data_only=True)
ws = wb[BAT_NAMES[BAT_ID]]
urls = [ws.cell(row=i + IPCAMS_DATABASE_STARTROW, column=IPCAMS_DATABASE_STARTCOL).value for i in
        range(1, ws.max_row + 1) if
        ws.cell(row=i + IPCAMS_DATABASE_STARTROW, column=IPCAMS_DATABASE_STARTCOL).value is not None]
print(urls)
n = len(urls)

# Prepare imgs
imgs = [np.zeros((H_DEFAULT, W_DEFAULT, 3)) for i in urls]

###########################################################

global img


def url_ok(url, timeout):
    try:
        r = requests.head(url, timeout=timeout)
        return r.status_code == 200
    except:
        return False


def update(index, url):
    # Read next stream frame in a daemon thread
    global w, h
    img_cv = None
    while True:
        success = False
        # Meth 1 for getting image from url
        try:
            if (url_ok(url, 1)):
                img_rawInput = urllib.request.urlopen(url)
                img_numpy = np.array(bytearray(img_rawInput.read()), dtype=np.uint8)
                img_cv = cv2.imdecode(img_numpy, -1)
                # Orientation correction
                img_cv = cv2.transpose(img_cv)
                if not are_frontals:
                    img_cv = cv2.flip(img_cv, 1)

                success = True
        except:
            print(f'failed to open cam n°{index:d}')

        # Meth 2 for getting image from url
        # try:
        #     if (url_ok(url, 1)):
        #         cap = cv2.VideoCapture(url)
        #         success, im = cap.read()
        # except:
        #     print(f'failed to open cam no : {index:d}')

        if img_cv is None:
            img_cv = np.zeros((h, w, 3))
        else:
            h, w = img_cv.shape[0], img_cv.shape[1]
        imgs[index] = img_cv
        #time.sleep(1 / 25)  # wait time


def detect(img_input):
    source, weights, view_img, save_txt, imgsz, trace = opt.source, opt.weights, opt.view_img, opt.save_txt, opt.img_size, not opt.no_trace

    # Initialize
    set_logging()
    device = select_device(opt.device)
    half = device.type != 'cpu'  # half precision only supported on CUDA

    # Load model
    model = attempt_load(weights, map_location=device)  # load FP32 model
    stride = int(model.stride.max())  # model stride
    imgsz = check_img_size(imgsz, s=stride)  # check img_size

    if trace:
        model = TracedModel(model, device, opt.img_size)

    if half:
        model.half()  # to FP16

    # Second-stage classifier
    classify = False
    if classify:
        modelc = load_classifier(name='resnet101', n=2)  # initialize
        modelc.load_state_dict(torch.load('weights/resnet101.pt', map_location=device)['model']).to(device).eval()

    view_img = check_imshow()
    cudnn.benchmark = True  # set True to speed up constant image size inference
    dataset = LoadStreams(source, img_size=imgsz, stride=stride)

    # Get names and colors
    names = model.module.names if hasattr(model, 'module') else model.names
    colors = [[random.randint(0, 255) for _ in range(3)] for _ in names]

    # Run inference
    if device.type != 'cpu':
        model(torch.zeros(1, 3, imgsz, imgsz).to(device).type_as(next(model.parameters())))  # run once
    old_img_w = old_img_h = imgsz
    old_img_b = 1

    t0 = time.time()
    img = torch.from_numpy(img_input).to(device)
    img = img.half() if half else img.float()  # uint8 to fp16/32
    img /= 255.0  # 0 - 255 to 0.0 - 1.0
    if img.ndimension() == 3:
        img = img.unsqueeze(0)

    # Warmup
    if device.type != 'cpu' and (
            old_img_b != img.shape[0] or old_img_h != img.shape[2] or old_img_w != img.shape[3]):
        old_img_b = img.shape[0]
        old_img_h = img.shape[2]
        old_img_w = img.shape[3]
        for i in range(3):
            model(img, augment=opt.augment)[0]

    # Inference
    t1 = time_synchronized()
    with torch.no_grad():  # Calculating gradients would cause a GPU memory leak
        pred = model(img, augment=opt.augment)[0]
    t2 = time_synchronized()

    # Apply NMS
    pred = non_max_suppression(pred, opt.conf_thres, opt.iou_thres, classes=opt.classes, agnostic=opt.agnostic_nms)
    t3 = time_synchronized()

    # Apply Classifier
    if classify:
        pred = apply_classifier(pred, modelc, img, im0s)

    # Process detections
    for i, det in enumerate(pred):  # detections per image
        if webcam:  # batch_size >= 1
            p, s, im0, frame = path[i], '%g: ' % i, im0s[i].copy(), dataset.count
        else:
            p, s, im0, frame = path, '', im0s, getattr(dataset, 'frame', 0)
        roi = im0  # For pose estimation
        p = Path(p)  # to Path

        gn = torch.tensor(im0.shape)[[1, 0, 1, 0]]  # normalization gain whwh
        if len(det):
            # Rescale boxes from img_size to im0 size
            det[:, :4] = scale_coords(img.shape[2:], det[:, :4], im0.shape).round()

            # Print results
            for c in det[:, -1].unique():
                n = (det[:, -1] == c).sum()  # detections per class
                s += f"{n} {names[int(c)]}{'s' * (n > 1)}, "  # add to string

            # Write results
            for *xyxy, conf, cls in reversed(det):
                if save_txt:  # Write to file
                    xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)) / gn).view(-1).tolist()  # normalized xywh
                    line = (cls, *xywh, conf) if opt.save_conf else (cls, *xywh)  # label format
                    with open(txt_path + '.txt', 'a') as f:
                        f.write(('%g ' * len(line)).rstrip() % line + '\n')

                if view_img:  # Add bbox to image
                    class_name = names[int(cls)]
                    label = f'{class_name} {conf:.2f}'
                    plot_one_box(xyxy, im0, label=label, color=colors[int(cls)], line_thickness=1)
                    xywh = (xyxy2xywh(torch.tensor(xyxy).view(1, 4)))

                    # Pose Estimation
                    if (class_name == "person"):
                        box_centerX = int(xywh[0][0])
                        box_centerY = int(xywh[0][1])
                        box_width = int(xywh[0][2])
                        box_height = int(xywh[0][3])
                        roi_tl = (box_centerX - int(box_width / 2), box_centerY - int(box_height / 2))
                        roi_br = (box_centerX + int(box_width / 2), box_centerY + int(box_height / 2))
                        roi = im0[roi_tl[1]: roi_br[1], roi_tl[0]: roi_br[0]]
                        if (poseEstimation_run(roi)):
                            poseEstimation_postProcess(im0, (roi_tl[0], roi_tl[1]))

        # Print time (inference + NMS)
        print(f'{s}Done. ({(1E3 * (t2 - t1)):.1f}ms) Inference, ({(1E3 * (t3 - t2)):.1f}ms) NMS')
    return im0



########## Prepare camera threads #########
for i, url in enumerate(urls):
    print(url)
    thread = Thread(target=update, args=([i, url]), daemon=True)
    thread.start()

########## Display live streams #########
while True:
    for i, img_cv_in in enumerate(imgs):
        if opt.update:  # update all models (to fix SourceChangeWarning)
            for opt.weights in ['yolov7.pt']:
                img_out = detect()
                strip_optimizer(opt.weights)
        else:
            img_out = detect()

        cv2.imshow(f'Live cam testing no : {i + 1:d}/{n:d} ({urls[i]})', img_out)
    key = cv2.waitKey(1)
    if key == ord('q'):
        break

cv2.destroyAllWindows()

# Initialize pose estimation
net = cv2.dnn.readNetFromTensorflow("graph_opt.pb")



with torch.no_grad():
    if opt.update:  # update all models (to fix SourceChangeWarning)
        for opt.weights in ['yolov7.pt']:
            detect()
            strip_optimizer(opt.weights)
    else:
        detect()